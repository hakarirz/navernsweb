import streamlit as st
from bs4 import BeautifulSoup
import requests
import openpyxl
import os
import datetime

def get_news_date(news_url): #뉴스의 날짜 확인 함수
    # 뉴스 기사 페이지에서 발행일 정보 추출
    try:
        r = requests.get(news_url, timeout=5) 
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
        date_element = soup.find("meta", {"property": "article:published_time"})
        if date_element:
            # 년, 월, 일만 추출
            date_content = date_element["content"]
            parsed_date = datetime.datetime.fromisoformat(date_content)
            return f"{parsed_date.year}년 {parsed_date.month}월 {parsed_date.day}일"
        else:
            return None  # 날짜가 없으면 None 반환
    except Exception as e:
        return None  # 에러가 발생하면 None 반환

def search_and_save_to_excel(keywords, num_news_per_keyword, start_date, end_date): #엑셀에 저장되는 함수
    # 파일 경로 설정
    folder_path = "results"
    os.makedirs(folder_path, exist_ok=True)  # 폴더가 없으면 생성
    fpath = os.path.join(folder_path, "모든키워드_뉴스.xlsx")

    # 기존 파일이 있다면 삭제하고 새로 생성
    if os.path.exists(fpath):
        os.remove(fpath)

    # 새로운 워크북 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    # 열 제목 추가
    ws.append(['키워드', '뉴스 번호', '뉴스 제목', '뉴스 링크', '뉴스 제작 날짜'])#5개로 설정

    for keyword in keywords:
        base_url = "https://search.naver.com/search.naver?where=news&sm=tab_jum&query=" #네이버 뉴스 1페이지
        search_url = f"{base_url}{keyword}"

        total_news_collected = 0
        page = 1 #페이지 1부터 시작

        while total_news_collected < num_news_per_keyword:
            r = requests.get(f"{search_url}&start={((page-1) * 10) + 1}") #페이지
            soup = BeautifulSoup(r.text, "html.parser")
            items = soup.select(".news_tit")

            for e, item in enumerate(items, start=1):
                # 뉴스 링크 가져오기
                news_link = item.get('href') if item.get('href') else '링크 없음' #링크 확인

                # 뉴스 제작 날짜 가져오기
                news_date = get_news_date(news_link)

                if news_date is not None:
                    # 날짜가 사용자가 지정한 날짜 범위에 속하면 추가
                    parsed_news_date = datetime.datetime.strptime(news_date, "%Y년 %m월 %d일").date()
                    if start_date <= parsed_news_date <= end_date:
                        # 엑셀에 데이터 추가
                        ws.append([keyword, e + total_news_collected, item.text, news_link, news_date])

                        if (e + total_news_collected) == num_news_per_keyword:
                            break

            total_news_collected += len(items)
            page += 1  # 페이지를 늘려야 네이버 뉴스의 다음 페이지를 갈 수 있음

    # 각 열의 최대 길이를 저장할 딕셔너리
    max_lengths = {'A': len('키워드'), 'B': len('뉴스 번호'), 'C': len('뉴스 제목'), 'D': len('뉴스 링크'), 'E': len('뉴스 제작 날짜')}

    # 기존 데이터의 최대 길이 업데이트
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        for col_num, value in enumerate(row, 1):
            max_lengths[chr(65 + col_num - 1)] = max(max_lengths[chr(65 + col_num - 1)], len(str(value)))

    # 각 열의 최대 길이를 기준으로 열의 너비를 동적으로 조절
    for column, max_length in max_lengths.items():
        adjusted_width = max_length + 18  # 추가 조절 가능
        ws.column_dimensions[column].width = adjusted_width

    # 엑셀 파일 저장
    wb.save(fpath)

    st.success(f"데이터가 {fpath}에 저장되었습니다.") 
    st.warning("키워드 검색시 파일은 초기화 됩니다.") 

# Streamlit 앱 생성
st.title('네이버 뉴스 크롤링')

# 텍스트 입력 및 뉴스 개수 입력 받기
user_input = st.text_input('키워드를 입력하세요 (여러 키워드는 쉼표로 구분)', '')
num_news_per_keyword = st.number_input('각 키워드 당 가져올 뉴스 개수를 입력하세요', min_value=1, value=5, step=1)
start_date = st.date_input('시작 날짜를 선택하세요', value=datetime.date(2023, 11, 28))
end_date = st.date_input('종료 날짜를 선택하세요', value=datetime.date(2023, 11, 29))

# 사용자 입력 표시 및 검색 결과 엑셀 파일 저장
if st.button('검색 및 저장'):
    if user_input:
        keywords = [kw.strip() for kw in user_input.split(',')] #공백 제거
        st.write(f'키워드: {", ".join(keywords)}를(을) 검색하셨습니다. 각 키워드 당 {num_news_per_keyword}개의 뉴스를 가져옵니다.')
        search_and_save_to_excel(keywords, num_news_per_keyword, start_date, end_date)
